# export_arc_length.py
from pyautocad import Autocad
import math
import openpyxl
from openpyxl import Workbook
import os
from collections import Counter

def get_arcs_from_layer(acad, doc, layer_name="外弧"):
    """
    从指定图层获取圆弧对象
    
    Args:
        acad: AutoCAD应用程序对象
        doc: AutoCAD文档对象
        layer_name: 图层名称，默认为"外弧"
    
    Returns:
        list: 指定图层中的圆弧对象列表
    """
    arcs = []
    
    try:
        # 遍历模型空间中的所有对象
        ms = doc.ModelSpace
        for i in range(ms.Count):
            try:
                entity = ms.Item(i)
                # 检查是否为圆弧且在指定图层上
                if entity.ObjectName == "AcDbArc" and entity.Layer == layer_name:
                    arcs.append(entity)
            except Exception as e:
                print(f"检查对象 {i} 时出错: {e}")
                continue
                
        print(f"从图层 '{layer_name}' 找到 {len(arcs)} 个圆弧对象")
        
    except Exception as e:
        print(f"获取图层对象时出错: {e}")
        
    return arcs

def calculate_arc_length(arc):
    """
    计算圆弧的长度并向上取整
    
    Args:
        arc: AutoCAD圆弧对象
    
    Returns:
        int: 向上取整后的圆弧长度
    """
    try:
        radius = arc.Radius
        start_angle = arc.StartAngle
        end_angle = arc.EndAngle
        
        # 计算圆弧角度差
        if end_angle < start_angle:
            angle_diff = (2 * math.pi - start_angle) + end_angle
        else:
            angle_diff = end_angle - start_angle
            
        # 计算弧长 = 半径 × 弧度
        arc_length = radius * angle_diff
        
        # 向上取整
        arc_length_ceil = math.ceil(arc_length)
        
        return arc_length_ceil
    except Exception as e:
        print(f"计算圆弧长度时出错: {e}")
        return 0

def export_arc_lengths_to_excel(arcs, filename="dxf_output/arc2excel/arc_lengths.xlsx"):
    """
    将圆弧长度导出到Excel文件（只包含取整后的弧长和数量）
    
    Args:
        arcs: 圆弧对象列表
        filename: 导出的Excel文件名
    """
    try:
        # 确保文件路径存在
        file_dir = os.path.dirname(filename)
        if file_dir and not os.path.exists(file_dir):
            os.makedirs(file_dir)
            print(f"创建目录: {file_dir}")
        
        # 计算所有圆弧的长度并向上取整
        arc_lengths = []
        for arc in arcs:
            length = calculate_arc_length(arc)
            arc_lengths.append(length)
        
        # 统计每种长度的数量
        length_counts = Counter(arc_lengths)
        
        # 创建工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "圆弧长度统计"
        
        # 写入表头
        ws['A1'] = "弧长(向上取整)"
        ws['B1'] = "数量"
        
        # 写入数据
        row = 2
        for length in sorted(length_counts.keys()):
            count = length_counts[length]
            ws[f'A{row}'] = length
            ws[f'B{row}'] = count
            print(f"弧长: {length}, 数量: {count}")
            row += 1
        
        # 保存文件
        wb.save(filename)
        print(f"成功导出弧长统计信息到 {filename}")
        
    except Exception as e:
        print(f"导出到Excel时出错: {e}")

def main():
    """
    主函数 - 导出"外弧"图层中圆弧的弧长到Excel（只包含取整后的弧长和数量）
    """
    # 连接到正在运行的 AutoCAD
    try:
        acad = Autocad(create_if_not_exists=True)
        doc = acad.doc
        print(f"成功连接到 AutoCAD 文档: {doc.Name}")
    except Exception as e:
        print("无法连接到 AutoCAD:", e)
        return
    
    try:
        # 从"外弧"图层获取圆弧对象
        arcs = get_arcs_from_layer(acad, doc, "外弧")
        
        if not arcs:
            print("在'外弧'图层中没有找到圆弧对象")
            return
        
        print(f"找到 {len(arcs)} 个圆弧对象")
        
        # 导出圆弧长度到Excel
        export_arc_lengths_to_excel(arcs, "dxf_output/arc2excel/外弧长度数据.xlsx")
            
    except Exception as e:
        print(f"处理对象时出错: {e}")

if __name__ == "__main__":
    main()