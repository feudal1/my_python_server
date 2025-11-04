# pdf_material_extractor_cli.py
import os
import json
from pathlib import Path
import fitz  # PyMuPDF
from PIL import Image
import pandas as pd
import time
import logging
from dotenv import load_dotenv
import http.client
from urllib.parse import urlparse
import base64
import io
from tkinter import Tk
from tkinter.filedialog import askdirectory, askopenfilenames  # 添加 askopenfilenames
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class PDFToImageConverter:
    @staticmethod
    def convert_pdf_to_pil_images(pdf_path):
        """
        将PDF文件转换为PIL图像对象列表（不保存到磁盘）
        
        Args:
            pdf_path (str): PDF文件路径
            
        Returns:
            list: PIL图像对象列表
        """
        logger.info(f"正在转换PDF为内存图像: {pdf_path}")
        
        try:
            # 打开PDF文件
            pdf_document = fitz.open(pdf_path)
            pil_images = []
            
            # 遍历每一页
            for page_num in range(len(pdf_document)):
                try:
                    # 获取页面
                    page = pdf_document[page_num]
                    
                    # 渲染页面为图像
                    mat = fitz.Matrix(2.0, 2.0)  # 2x zoom for better quality
                    pix = page.get_pixmap(matrix=mat)
                    
                    # 直接转换为PIL图像而不保存到磁盘
                    img_data = pix.tobytes("ppm")
                    pil_image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    pil_images.append(pil_image)
                    
                    logger.info(f"已转换PDF第 {page_num + 1} 页为内存图像")
                except Exception as page_error:
                    logger.error(f"处理PDF第 {page_num + 1} 页时出错: {str(page_error)}")
                    continue  # 继续处理其他页面
            
            pdf_document.close()
            return pil_images
        except Exception as e:
            logger.error(f"打开PDF文件 {pdf_path} 时出错: {str(e)}")
            return []

class VLMProcessor:
    def __init__(self):
        """初始化VLM处理器"""
        dotenv_path = r'E:\code\apikey\.env'
        load_dotenv(dotenv_path)
        self.api_url = os.getenv('VLM_OPENAI_API_URL')
        self.model_name = os.getenv('VLM_MODEL_NAME')
        self.api_key = os.getenv('VLM_OPENAI_API_KEY')
        
        # 定义提取字段配置
        self.extraction_fields = [
            {"key": "零件", "label": "零件1", "display_name": "零件名称"},
            {"key": "材料", "label": "Q235B", "display_name": "材料"},
            {"key": "数量", "label": "1", "display_name": "数量"},
            {"key": "文件名", "label": "文件1.dwg", "display_name": "零件文件名"}
        ]
        
        if not all([self.api_url, self.model_name, self.api_key]):
            raise ValueError("缺少VLM API配置，请检查.env文件")
    
    def encode_pil_image(self, pil_image):
        """
        将PIL图像编码为base64
        
        Args:
            pil_image (PIL.Image): PIL图像对象
            
        Returns:
            str: base64编码的图像
        """
        # 将PIL图像保存到内存中的字节流
        buffer = io.BytesIO()
        pil_image.save(buffer, format='PNG')
        # 编码为base64
        return base64.b64encode(buffer.getvalue()).decode('utf-8')
    
    def extract_material_info(self, pil_image):
        """
        使用VLM提取材料信息
        
        Args:
            pil_image (PIL.Image): PIL图像对象
            
        Returns:
            dict: 材料信息
        """
        logger.info("正在提取材料信息")
        
        # 编码图像
        base64_image = self.encode_pil_image(pil_image)
        
        # 根据当前字段配置动态生成提示语句
        labels = [field["display_name"] for field in self.extraction_fields]
        labels_str = "、".join(labels)
        instruction = f"请从这张图纸中提取零件信息，包括{labels_str}等信息"
        
        # 动态构建JSON示例中的字段部分
        field_lines = []
        for field in self.extraction_fields:
            field_lines.append(f'            "{field["key"]}": "{field["label"]}"')
        
        json_fields_block = ",\n".join(field_lines)
        
        # 构造消息
        messages = [
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": f'''{instruction}。如果有多个零件，请分别列出。以JSON数组格式返回。
输出格式要求：
[
    {{
{json_fields_block}
    }}
]

如果无法识别，请返回[{{"零件": "未知", "材料": "未知", "数量": "未知", "文件名": "未知"}}]'''
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/png;base64,{base64_image}"
                        }
                    }
                ]
            }
        ]
        
        return self._send_request(messages)
    
    def _send_request(self, messages):
        """
        发送请求到VLM API
        
        Args:
            messages (list): 消息列表
            
        Returns:
            dict: API响应结果
        """
        # 解析URL
        parsed = urlparse(f"{self.api_url}/chat/completions")
        host, path = parsed.hostname, parsed.path
        
        # 创建连接
        conn = http.client.HTTPSConnection(host)
        
        # 构造请求体
        request_body = {
            "model": self.model_name,
            "messages": messages,
            "temperature": 0.7
        }
        
        # 发送请求
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {self.api_key}"
        }
        
        conn.request(
            "POST",
            path,
            body=json.dumps(request_body),
            headers=headers
        )
        
        # 获取响应
        response = conn.getresponse()
        
        if response.status != 200:
            error_msg = response.read().decode('utf-8')
            conn.close()
            raise Exception(f"VLM服务器错误: {response.status} - {error_msg}")
        
        # 解析响应
        response_data = response.read().decode('utf-8')
        data = json.loads(response_data)
        conn.close()
        
        # 提取结果
        try:
            content = data['choices'][0]['message']['content']
            # 处理可能的Markdown包装
            if content.startswith("```json"):
                content = content[7:]  # 移除 ```json
                if content.endswith("```"):
                    content = content[:-3]  # 移除 ```
            
            result = json.loads(content)
            logger.info(f"VLM处理结果: {result}")
            return result
        except Exception as e:
            logger.error(f"解析VLM响应失败: {e}")
            logger.error(f"原始响应: {content}")
            raise

class ExcelUpdater:
    def __init__(self):
        # 定义提取字段配置
        self.extraction_fields = [
            {"key": "零件", "display_name": "零件名称"},
            {"key": "材料", "display_name": "材料"},
            {"key": "数量", "display_name": "数量"},
            {"key": "文件名", "display_name": "零件文件名"}
        ]
    
    def _convert_to_number_if_possible(self, value):
        """
        如果可能，将字符串转换为数字（int或float）
        """
        if isinstance(value, str):
            # 去除首尾空格
            value = value.strip()
            
            # 尝试转换为整数
            if value.isdigit() or (value.startswith('-') and value[1:].isdigit()):
                return int(value)
            
            # 处理Unicode数字字符（如①②③等）
            try:
                unicode_digits = {
                    '①': 1, '②': 2, '③': 3, '④': 4, '⑤': 5,
                    '⑥': 6, '⑦': 7, '⑧': 8, '⑨': 9, '⑩': 10
                }
                if value in unicode_digits:
                    return unicode_digits[value]
            except:
                pass
            
            # 尝试转换为浮点数
            try:
                # 处理可能包含逗号的数字（如"1,234.56"）
                value_normalized = value.replace(',', '')
                if '.' in value_normalized:
                    float_val = float(value_normalized)
                    # 如果小数部分为0，则返回整数
                    if float_val.is_integer():
                        return int(float_val)
                    else:
                        return float_val
            except ValueError:
                pass
        
        # 如果不能转换为数字，返回原始值
        return value
    
    def update_excel(self, excel_path, file_name, material_info_list):
        """
        更新Excel文件中的信息
        
        Args:
            excel_path (str): Excel文件路径
            file_name (str): PDF文件名(不含扩展名)
            material_info_list (list): 材料信息列表
        """
        logger.info(f"正在更新Excel: {excel_path}")
        
        # 确保 material_info_list 是一个列表
        if not isinstance(material_info_list, list):
            material_info_list = [material_info_list]
        
        # 创建工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "材料信息"
        
        # 在A1单元格写入文件名并合并前五行
        ws["A1"] = f"PDF文件名: {file_name}"
        ws.merge_cells("A1:D5")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].font = Font(size=14, bold=True)
        
        # 设置标题行行高
        for row_num in range(1, 6):
            ws.row_dimensions[row_num].height = 25
        
        # 创建显示名称映射
        display_names = {field["key"]: field["display_name"] for field in self.extraction_fields}
        
        # 创建表头行（中文）
        header_row = ["PDF文件名"]
        field_order = ["PDF文件名"]  # 保持列顺序
        
        # 按预定义顺序添加字段
        for field_config in self.extraction_fields:
            field_key = field_config["key"]
            header_row.append(display_names.get(field_key, field_key))
            field_order.append(field_key)
        
        # 写入表头
        for col_num, header in enumerate(header_row, 1):
            cell = ws.cell(row=6, column=col_num, value=header)
            cell.font = Font(bold=True)
        
        # 填充数据行
        start_row = 7
        for i, material_info in enumerate(material_info_list):
            row_num = start_row + i
            # PDF文件名列
            ws.cell(row=row_num, column=1, value=file_name)
            
            # 其他字段列
            for col_idx, field in enumerate(self.extraction_fields, 2):  # 从第2列开始
                cell_value = material_info.get(field["key"], "未知")
                # 尝试将值转换为数字
                final_value = self._convert_to_number_if_possible(cell_value)
                ws.cell(row=row_num, column=col_idx, value=final_value)
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # 限制最大宽度
        
        # 保存Excel文件
        try:
            wb.save(excel_path)
            logger.info(f"已成功更新Excel文件，保存路径: {excel_path}")
        except Exception as e:
            logger.error(f"保存Excel文件失败: {str(e)}")
            raise
        logger.info("已更新Excel文件")

def select_pdf_files():
    """打开文件选择对话框，选择多个PDF文件"""
    try:
        # 隐藏主窗口
        root = Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        
        files = askopenfilenames(
            title="选择PDF文件",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        root.destroy()
        
        if files:
            # 只保留PDF文件
            pdf_files = [f for f in files if f.lower().endswith('.pdf')]
            return pdf_files
        else:
            return []
    except Exception as e:
        logger.error(f"选择文件时出错: {str(e)}")
        return []

def main():
    """主函数"""
    print("启动PDF材料信息提取工具...")
    
    # 选择PDF文件
    pdf_files = select_pdf_files()
    
    if not pdf_files:
        print("未选择任何PDF文件，程序退出。")
        return
    
    # 验证所有文件是否存在
    for pdf_file in pdf_files:
        if not os.path.exists(pdf_file):
            print(f'PDF文件不存在: {pdf_file}')
            return
    
    # 设置Excel输出路径为第一个PDF文件所在的目录
    first_pdf_dir = os.path.dirname(pdf_files[0])
    excel_path = os.path.join(first_pdf_dir, 'pdf_material_info.xlsx')
    
    print(f"已选择 {len(pdf_files)} 个PDF文件:")
    for i, pdf_file in enumerate(pdf_files):
        print(f"  {i+1}. {os.path.basename(pdf_file)}")
    print(f"Excel文件将保存至: {excel_path}")
    
    # 初始化组件
    converter = PDFToImageConverter()
    processor = VLMProcessor()
    updater = ExcelUpdater()
    
    # 缓存
    pil_images_cache = {}
    
    # 处理每个PDF文件
    for index, pdf_file in enumerate(pdf_files):
        print(f"\n开始处理PDF文件 ({index+1}/{len(pdf_files)}): {os.path.basename(pdf_file)}")
        
        try:
            # 检查是否已经转换过该PDF
            pdf_name = Path(pdf_file).stem
            if pdf_name in pil_images_cache:
                # 如果已经转换过，直接使用已有的图像对象
                pil_images = pil_images_cache[pdf_name]
                print(f"使用已缓存的图像对象，共 {len(pil_images)} 页")
            else:
                # 转换PDF为图像对象
                pil_images = converter.convert_pdf_to_pil_images(pdf_file)
                
                # 缓存图像对象
                pil_images_cache[pdf_name] = pil_images
                print(f"新生成图像对象并缓存，共 {len(pil_images)} 页")
            
            material_info_list = None
            
            # 使用第一页作为预览
            if pil_images and len(pil_images) > 0:
                first_pil_image = pil_images[0]
                print("正在提取材料信息...")
                
                # 提取材料信息
                try:
                    material_result = processor.extract_material_info(first_pil_image)
                    print(f"材料信息提取结果: {material_result}")
                    
                    # 确保结果是列表格式
                    if isinstance(material_result, dict):
                        material_info_list = [material_result]
                    elif isinstance(material_result, list):
                        material_info_list = material_result
                    else:
                        material_info_list = [{"零件": "未知", "材料": "提取失败", "数量": "提取失败", "文件名": "提取失败"}]
                        
                except Exception as e:
                    print(f"材料信息提取失败: {str(e)}")
                    material_info_list = [{"零件": "未知", "材料": "提取失败", "数量": "提取失败", "文件名": "提取失败"}]
                
                # 更新Excel
                try:
                    updater.update_excel(excel_path, pdf_name, material_info_list)
                    print("Excel更新完成")
                except Exception as e:
                    print(f"更新Excel时出错: {str(e)}")
            else:
                print(f"未生成图像对象 for {pdf_file}")
                
        except Exception as e:
            print(f"处理文件 {pdf_file} 时出错: {str(e)}")
    
    print("\n所有文件处理完成！")
    print(f"Excel文件已保存至: {excel_path}")

if __name__ == "__main__":
    main()