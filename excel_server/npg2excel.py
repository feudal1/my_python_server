# vlm_png_to_excel.py
import json
import os
from PIL import Image
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import tempfile
from flask import Flask, request, jsonify, send_file
from werkzeug.utils import secure_filename
# 添加路径设置
import sys
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

# 修改导入语句
from llm_server.llm_server import VLMService

# Flask 应用实例
app = Flask(__name__)

# 创建 VLMService 实例
vlm_service = VLMService()

def parse_vlm_response(response_data):
    """
    解析 VLM 返回的结构化数据
    """
    try:
        # 获取响应内容
        result = response_data.get("choices", [{}])[0].get("message", {}).get("content", "{}")
        
        # 去除可能的代码块标记
        if result.startswith("```json"):
            result = result[7:]  # 去除开头的 ```json
        if result.startswith("```"):
            result = result[3:]  # 去除可能的 ```
        if result.endswith("```"):
            result = result[:-3]  # 去除结尾的 ```
        
        # 去除首尾空白字符
        result = result.strip()
        
        # 解析JSON
        parsed = json.loads(result)
        
        # 只保留 drawing_info 字段
        if "drawing_info" in parsed:
            return {"drawing_info": parsed["drawing_info"]}
        else:
            return {}
    except Exception as e:
        print(f"解析失败: {e}")
        return {}

def save_data_to_excel(data, output_path="output.xlsx", image_filename=None):
    """
    将结构化数据保存为Excel文件
    
    Args:
        data (dict): 结构化数据
        output_path (str): 输出Excel文件路径
        image_filename (str): 原始图像文件名，用于填入Excel前三行合并单元格
    """
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    
    # 如果提供了图像文件名，在前三行合并单元格并填入文件名
    if image_filename:
        # 写入文件名到第一行
        ws["A1"] = image_filename
        # 合并前三行
        ws.merge_cells("A1:E3")
        # 设置居中对齐和字体
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].font = Font(size=14, bold=True)
        
        # 从第4行开始写入数据标题
        start_row = 4
    else:
        start_row = 1
    
    # 准备数据
    rows = []
    
    # 处理图纸基本信息
    if "drawing_info" in data and data["drawing_info"]:
        rows.append(["图纸信息", "", "", "", ""])
        rows.append(["图号", "长度(mm)", "宽度(mm)", "数量", "备注"])
        for item in data["drawing_info"]:
            drawing_number = item.get("drawing_number", "")
            length = item.get("length", "")
            width = item.get("width", "")
            quantity = item.get("quantity", "")
            remark = item.get("remark", "")
            rows.append([drawing_number, length, width, quantity, remark])
    
    # 移除了 dimensions 和 text 的处理部分
    
    # 将数据写入工作表
    for i, row in enumerate(rows, start=start_row):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存文件
    wb.save(output_path)
    print(f"Excel 文件已保存至: {output_path}")
    return output_path

def analyze_image_and_generate_excel(image_path):
    """
    主函数：调用 VLM 分析图像，并生成 Excel 文件
    
    Args:
        image_path (str): 输入图像路径
        output_path (str): 输出Excel文件路径，默认为"dxf_output/data/extracted_data.xlsx"
    """
   
    output_path = r"excel_output/npg2excel/extracted_data.xlsx"
    
    # 确保输出目录存在
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # Step 1: 构造请求消息
    messages = [
        {
            "role": "user",
            "content": '''请提取这张图片的信息，尽可能提取出图号、数量（要算）等信息，图号类似F2N-数字并以 JSON 格式返回。
输出格式要求：
{
    "drawing_info": [
        {
            "drawing_number": "ABC123",
            "length": "1200",
            "width": "800",
            "quantity": "5",
            "remark": "不锈钢板"
        }
    ]
}'''
        }
    ]

    # Step 2: 直接调用 VLMService 而不是通过 HTTP 请求
    try:
        result = vlm_service.create_with_image(messages, image_path)
        print("VLM 返回:", result)
        # 解析响应
        extracted_data = parse_vlm_response(result)
        # 生成 Excel
        save_data_to_excel(extracted_data, output_path, os.path.basename(image_path))
    except Exception as e:
        print("错误:", str(e))

@app.route('/api/analyze-image', methods=['GET'])
def analyze_image_api():
    """
    Flask 接口：通过URL参数指定图片路径，调用VLM分析并生成Excel文件
    
    参数:
    - image_path: 图片路径 (必需)

    """
    try:
        # 从请求参数中获取图片路径
        image_path = request.args.get('image_path')
     
         
        if not image_path:
            return jsonify({"error": "缺少 image_path 参数"}), 400
        
        # 检查输入文件是否存在
        if not os.path.exists(image_path):
            return jsonify({"error": f"图像文件 '{image_path}' 不存在"}), 400
            
        # 创建临时文件保存输出的Excel
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_excel:
            output_path = tmp_excel.name
        
        # 调用分析函数
        analyze_image_and_generate_excel(image_path)
        
        # 返回生成的Excel文件
        return send_file(
            output_path,
            as_attachment=True,
           
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/routes')
def show_routes():
    routes = []
    for rule in app.url_map.iter_rules():
        # 获取端点函数的docstring
        endpoint_func = app.view_functions.get(rule.endpoint)
        docstring = endpoint_func.__doc__.strip() if endpoint_func and endpoint_func.__doc__ else "无描述"
        
        routes.append({
            'endpoint': rule.endpoint,
            'methods': list(rule.methods),
            'rule': str(rule),
            'description': docstring
        })
    return {'routes': routes}

if __name__ == '__main__':
    # 检查是否以脚本方式直接运行
    if len(sys.argv) == 1:
        # 使用 input() 获取图像路径
        print("请输入图像文件路径（输入 'quit' 退出）:")
        while True:
            image_path = input("图像路径: ").strip()
            
            if image_path.lower() == 'quit':
                print("程序退出")
                break
                
            if not os.path.exists(image_path):
                print(f"错误: 图像文件 '{image_path}' 不存在，请重新输入")
                continue
                
            try:
                analyze_image_and_generate_excel(image_path)
                print("处理完成！")
                break
            except Exception as e:
                print(f"处理过程中出现错误: {e}")
                continue
    else:
        # Flask 服务器模式
        import argparse
        parser = argparse.ArgumentParser()
        parser.add_argument('--port', type=int, default=5001, help='Port to run the server on')
        args = parser.parse_args()
        app.run(host='0.0.0.0', port=args.port, debug=True)